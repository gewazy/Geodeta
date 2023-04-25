import pyodbc
import openpyxl
from datetime import date, timedelta, datetime
from time import sleep
import kury
import init

# do zrobienia:
# - rozpoznie ostatniego dnia w raporcie i przygotowanie xlsx do dnia obecnego, oraz wypełnienie dni po kolei
#   /jesli oststnia zakladka jest starsza niż z dnia dzisiejszego iterowac przez skrypt dla poszczegolnych dni


print('Generowania raportu DPR.\n\tver4.beta')

report_day = str((date.today() - timedelta(days=init.DDIFF)).strftime('%Y%m%d'))
print(f"-dane z dnia: {report_day}")

sleep(0.5)

# łączenie z bazą
cnxn = pyodbc.connect(init.CONN_STR)  # łączenie z bazą danych
crsr = cnxn.cursor()

# zbieranie danch do wypełnienia raportu
vib = len(crsr.execute(kury.VIB).fetchall())
xr = len(crsr.execute(kury.XR).fetchall())
xt = len(crsr.execute(kury.XT).fetchall())
skip = len(crsr.execute(kury.SKIP_S).fetchall())
qc_r = len(crsr.execute(kury.QC_R).fetchall())
qc_s = len(crsr.execute(kury.QC_S).fetchall())
paterny = len(crsr.execute(kury.PATERNY).fetchall())


tycz_r = crsr.execute(kury.TYCZ_R).fetchall()
tycz_s = crsr.execute(kury.TYCZ_S).fetchall()

zm_r = crsr.execute(kury.ZM_R).fetchall()
zm_s = crsr.execute(kury.ZM_S).fetchall()

re_s = crsr.execute(kury.RE_S).fetchall()
re_r = crsr.execute(kury.ZM_R).fetchall()
try:
    otg = crsr.execute(kury.OTG).fetchall()
except pyodbc.ProgrammingError:
    print('Brak tabeli OTG')

crsr.close()
cnxn.close()


# Przygotowanie pliku Excell
print('Otwieram DPR\n')

wb = openpyxl.load_workbook(init.DPR[0])

print('-Tworzę nową zakładkę')
source = wb[wb.sheetnames[-2]]
target = wb.copy_worksheet(source)
wb.move_sheet(target, offset=-1)
sleep(1)
print('-Zmieniam nazwe zakładki')
target.title = str((datetime.strptime(report_day, '%Y%m%d') + timedelta(days=1)).strftime('%Y%m%d'))
target.sheet_view.zoomScale = 75
target.sheet_view.view = "pageBreakPreview"
sleep(1)
# zamiana formuł w pliku
print('-Edytuję formuły')
a = f"'{str((datetime.strptime(source.title, '%Y%m%d') - timedelta(days=1)).strftime('%Y%m%d'))}'"
b = f"'{source.title}'"
for row in target.iter_rows(min_row=1, max_col=15, max_row=133):
    for cell in row:
        if type(cell.value) is str:
            cell.value = cell.value.replace(a, b)
sleep(1)

# ustawienie arkusza aktywnego
print('\nPrzechodzę na arkusz z dnia raportowanego')
target = wb[wb.sheetnames[-3]]
sleep(1)

# wypełnianie raportu
print("-Wypełniam arkusz\n")
sleep(1)

print('--Wibratory: ', vib)
target['E132'] = vib
sleep(1)

print('--Wiercenie ręczne: ', xr)
target['F132'] = xr
sleep(1)

print('--Wiercenie traktorem: ', xt)
target['G132'] = xt
sleep(1)

print('--W tym patterny: ', paterny)
target['H134'] = paterny
sleep(1)

print('--Skipy: ', skip)
target['K132'] = skip
sleep(1)

print('--QC R: ', qc_r)
if qc_r != 0:
    target['G53'] = qc_r
sleep(1)

print('--QC S: ', qc_s)
if qc_s != 0:
    target['O53'] = qc_s
sleep(1)


licz_bry = []  # lista brygadzistów do liczenia liczby brygad

row = 13
print('\nTyczenie punktów odbioru: \n')
for ro in tycz_r:
    row += 1
    print(ro)
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    target['D' + str(row)] = ro[3]
    try:
        if (ro[2].split())[1] != 'SKIP':  # wyklucza skipy z liczenia brygadzistow
            licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue
sleep(1)

row = 13
print('\nTyczenie punktów wzbudzania: \n')
for ro in tycz_s:
    row += 1
    print(ro)
    target['I' + str(row)] = ro[0]
    target['J' + str(row)] = ro[1]
    target['K' + str(row)] = ro[2]
    target['L' + str(row)] = ro[3]
    try:
        if ((ro[2].split()))[1] != 'SKIP':  # wyklucza skipy z liczenia brygadzistow
            licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue
sleep(1)

row = 41
print('\nDomierzanie/niwelacja punktów odbioru: \n')
for ro in re_r:
    row += 1
    print(ro)
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    target['D' + str(row)] = ro[3]
    try:
        licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue
sleep(1)

row = 41
print('\nDomierzanie/niwelacja punktów wzbudzania: \n')
for ro in re_s:
    row += 1
    print(ro)
    target['I' + str(row)] = ro[0]
    target['J' + str(row)] = ro[1]
    target['K' + str(row)] = ro[2]
    target['L' + str(row)] = ro[3]
    try:
        licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue
sleep(1)

# rozróżnienie zmian od niwelacji/domierzania /geofony/
for r1 in zm_r:
    for r2 in re_r:
        if r1[2] == r2[2]:
            r1[3] -= r2[3]

row = 57
print('\nZmiany punktów odbioru: \n')
for ro in zm_r:
    if ro[3] > 0:
        row += 1
        print(ro)
        target['A' + str(row)] = ro[0]
        target['B' + str(row)] = ro[1]
        target['C' + str(row)] = ro[2]
        target['D' + str(row)] = ro[3]
        try:
            licz_bry.append((ro[2].split())[1])
        except IndexError:
            continue
sleep(1)

# rozróżnienie zmian od niwelacji, domierzania /pkt strzałowe/
for r1 in zm_s:
    for r2 in re_s:
        if r1[2] == r2[2]:
            r1[3] -= r2[3]

row = 57
print('\nZmiany punktów wzbudzania: \n')
for ro in zm_s:
    if ro[3] > 0:
        row += 1
        print(ro)
        target['I' + str(row)] = ro[0]
        target['J' + str(row)] = ro[1]
        target['K' + str(row)] = ro[2]
        target['L' + str(row)] = ro[3]
        try:
            licz_bry.append((ro[2].split())[1])
        except IndexError:
            continue
sleep(1)

row = 95
print('\nPomiar otworów głębokich: \n')
try:
    for ro in otg:
        row += 1
        print(ro)
        target['A' + str(row)] = ro[0]
        target['B' + str(row)] = ro[1]
        target['C' + str(row)] = ro[2]
        target['D' + str(row)] = ro[3]
        try:
            licz_bry.append((ro[2].split())[1])
        except IndexError:
            continue
except NameError:
    print('Brak pomiarów OTG')

sleep(1)

print(f'\n\nPracowało {len(set(licz_bry))} brygad\n')
for num, geodeta in enumerate(sorted(set(licz_bry))):
    print(str(num + 1) + '.', geodeta)
sleep(1)

target['N132'] = len(set(licz_bry))
target['B128'] = init.comment(init.KOMENTARZ)

print('\nZapisuję plik')
wb.save(init.DPR[0])
sleep(1)

input('\nEnter by zakończyć!')
