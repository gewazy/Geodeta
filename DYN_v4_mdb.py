import csv
import datetime
import openpyxl_dictreader
import openpyxl
import pyodbc
import pyproj
from tkinter import Tk
from tkinter.filedialog import askopenfilename

# v4 skrypt sam nadaje także datę juliańską. DJ
import init


def tran(lat, lon):
    transformer = pyproj.Transformer.from_crs(4326, 2180)
    return transformer.transform(lat, lon)


print("DYNAMIT")

Tk().withdraw()  # otworzenie okna
open_filename = askopenfilename()  # wskazanie pliku do otwarcia
print("Otworz plik: " + open_filename)

save_filename = open_filename[:-5] + '_OK.csv' # info o pliku zapisywanym
print("Zapisz plik: " + save_filename)


# praca z excelem
wb = openpyxl.load_workbook(open_filename)
sheets = wb.sheetnames
print('\nw pliku BoomBox znalazłem następujące zakładki: \n' + str(sheets))

ws = wb.sheetnames[0]
print('pracuję na zakładce: \n' + ws)

ws2 = wb[ws]

#  liczenie daty juliańskiej
jd = datetime.datetime(int(ws[8:12]), int(ws[13] if ws[12] == '0' else ws[12:14]),
                       int(ws[15] if ws[14] == '0' else ws[14:16])).strftime('%Y%j')

input(f'\nJedziemy? \nDzień juliański: {jd}, Ok? -> Enter ')

ws2.delete_cols(17)
wb.save(open_filename)

# .....
dynamitxls = openpyxl_dictreader.DictReader(open_filename, ws)



numline = ws2.max_row - 1

with open(save_filename, 'w', newline='') as w_DYN:

    fieldnames = dynamitxls.fieldnames + ['Name', 'JulianDay', 'Local Easting', 'Local Northing']
    writer_csv = csv.DictWriter(w_DYN, fieldnames)
    writer_csv.writeheader()

    prog = 0
    for row in dynamitxls:
        prog += 1
        row['Name'] = str(row['Line']) + str(row['Station'])
        row['JulianDay'] = jd
        row['Local Northing'], row['Local Easting'] = tran(float(row['Lat(deg N)']), float(row['Lon(deg E)']))
        writer_csv.writerow(row)
        print(f"\rPostęp {prog} z {numline}; Procentowo {int(prog * 100 / numline)} %", end="")

input('\nENTER żeby wgrać do bazy danych; zamknij okno żeby wyjść')

params = []

with open(save_filename, 'r') as r_DYN:
    dict_reader_DYN = csv.DictReader(r_DYN)
    for row in dict_reader_DYN:
        params.append((row['Crew'], row['Unit'], row['ID'], row['Line'], row['Station'], row['Time'], row['Status'],
                       row['First Pick(ms)'], row['CTB(us)'], row['Uphole Shift(ms)'], row['Cap res(ohm)'],
                       row['Geo res(ohm)'], row['Battery(V)'], row['Lat(deg N)'], row['Lon(deg E)'], row['Alt(m)'],
                       row['Name'], row['JulianDay'], row['Local Easting'], row['Local Northing']))

for row in params:
    print(row[3], row[4], row[-3], round(float(row[-2]), 2), round(float(row[-1]), 2))

# łaczenie z bazą danych i wgranie danych
cnxn = pyodbc.connect(init.CONN_STR)  # łączenie z bazą danych
crsr = cnxn.cursor()

crsr.executemany("INSERT INTO DYN ([Crew], [Unit], [ID], [Line], [Station], [Time], [Status], [First Pick(ms)],"
                 " [CTB(us)], [Uphole Shift(ms)], [Cap res(ohm)], [Geo res(ohm)], [Battery(V)],"
                 " [Lat(deg N)], [Lon(deg E)], [Alt(m)], [Name], [JulianDay], [Local Easting], [Local Northing])"
                 "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", params)
crsr.close()
cnxn.commit()
cnxn.close()


input('zrobione!! \n\nENTER żeby zamknąć okno')