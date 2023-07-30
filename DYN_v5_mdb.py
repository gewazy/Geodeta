import csv
import datetime
import openpyxl_dictreader
import openpyxl
import pyodbc
import pyproj
from tkinter import Tk
from tkinter.filedialog import askopenfilename
import re

import init

# Zamiany 30-07-2023 djurkowski:
# - inny sposób szukania daty - wcześniej z nazwy zakładki obecnie za pomocą modułu .re z znazwy pliku (openfilename).
#   uniezależniamy się od róznych nazw zakładek (sejsmolog na pl192 używa sryptu który nadaje inną nazwę zakłądki).
# - bardziej czytelne liczenie daty juliańskiej


def tran(lat, lon):
    transformer = pyproj.Transformer.from_crs(4326, 2180)
    return transformer.transform(lat, lon)


print("DYNAMIT")

Tk().withdraw()  # otworzenie okna
open_filename = askopenfilename()  # wskazanie pliku do otwarcia
# szukanie daty w nazwie pliku.
np = re.search(r'\d{8}', open_filename, flags=re.IGNORECASE).group()
d, m, y = int(np[:4]), int(np[4:6]), int(np[6:])

print("Otworz plik: " + open_filename)

save_filename = open_filename[:-5] + '_OK.csv' # info o pliku zapisywanym
print("Zapisz plik: " + save_filename)


# praca z excelem
wb = openpyxl.load_workbook(open_filename)
sheets = wb.sheetnames
print('\nw pliku BoomBox znalazłem następujące zakładki: \n' + str(sheets))

ws = wb.sheetnames[0]
print('pracuję na zakładce: ' + ws)
print(f'Czas generowania danych {d}-{m}-{y}')

ws2 = wb[ws]

#  liczenie daty juliańskiej - dj
dzien = datetime.datetime(d, m, y)  # potrzebne w wypełnianiu pliku *.csv
jd = int(dzien.strftime('%Y%j'))


input(f'\nJedziemy? \nDzień juliański: {jd}, Ok? -> Enter ')

ws2.delete_cols(17)
wb.save(open_filename)

# .....
dynamitxls = openpyxl_dictreader.DictReader(open_filename, ws)


numline = ws2.max_row - 1
print(numline)

with open(save_filename, 'w', newline='') as w_DYN:

    fieldnames = dynamitxls.fieldnames + ['Name', 'JulianDay', 'Local Easting', 'Local Northing']
    writer_csv = csv.DictWriter(w_DYN, fieldnames)
    writer_csv.writeheader()

    prog = 0
    for row in dynamitxls:
        prog += 1
        row['Name'] = str(row['Line']) + str(row['Station'])
        row['JulianDay'] = jd
        try:
            row['Local Northing'], row['Local Easting'] = tran(float(row['Lat(deg N)']), float(row['Lon(deg E)']))
        except TypeError:
            row['Local Northing'], row['Local Easting'] = (0, 0)
        writer_csv.writerow(row)
        print(f"\rPostęp {prog} z {numline}; Procentowo {int(prog * 100 / numline)} %", end="")

input('\nENTER żeby wgrać do bazy danych; zamknij okno żeby wyjść')

params = []

# uwzględnia także błąd operatora boomboxa - brak pomiaru współrzędnych (wszystkie else). dj
with open(save_filename, 'r') as r_DYN:
    dict_reader_DYN = csv.DictReader(r_DYN)
    for row in dict_reader_DYN:
        params.append(
                      (row['Crew'] if row['Crew'] else '',
                       row['Unit'] if row['Unit'] else '',
                       row['ID'],
                       row['Line'] if row['Line'] else 0,
                       row['Station'] if row['Station'] else 0,
                       row['Time'] if row['Time'] else dzien,  # wstawienie daty generowania raportu
                       row['Status'] if row['Status'] else '',
                       row['Lat(deg N)'] if row['Lat(deg N)'] else 0,
                       row['Lon(deg E)'] if row['Lon(deg E)'] else 0,
                       row['Alt(m)'] if row['Alt(m)'] else 0,
                       row['Name'] if row['Name'] else 0,
                       row['JulianDay'],
                       row['Local Easting'],
                       row['Local Northing'])
                     )

for row in params:
    print(row[3], row[4], row[-3], round(float(row[-2]), 2), round(float(row[-1]), 2))

# łaczenie z bazą danych i wgranie danych
cnxn = pyodbc.connect(init.CONN_STR)  # łączenie z bazą danych
crsr = cnxn.cursor()

try:
    crsr.execute('select * from DYN;')
except pyodbc.ProgrammingError:
    crsr.execute(
                 "CREATE TABLE DYN ("
                 "[Crew] text(32),"
                 "[Unit] text(32),"
                 "[ID] text(32),"
                 "[Line] Long,"
                 "[Station] Long,"
                 "[Time] Date,"
                 "[Status] text(32),"
                 "[Lat(deg N)] Double, "
                 "[Lon(deg E)] Double, "
                 "[Alt(m)] Single,"
                 "[Name] Long,"
                 "[JulianDay] Long, "
                 "[Local Easting] Double, "
                 "[Local Northing] Double);")

crsr.executemany("INSERT INTO DYN ([Crew], [Unit], [ID], [Line], [Station], [Time], [Status], "
                 "[Lat(deg N)], [Lon(deg E)], [Alt(m)], [Name], [JulianDay], [Local Easting], [Local Northing])"
                 "VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?);", params)

crsr.close()
cnxn.commit()
cnxn.close()


input('zrobione!! \n\nENTER żeby zamknąć okno')
