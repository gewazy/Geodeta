import openpyxl
import pyodbc
from datetime import date, timedelta
from time import sleep

'''Wstawianie danych do raportu ze statystykami'''
'''Poki co problem z xlsm - szukam rozwiazania'''

print("Tworzę połaczenie z bazą danych")
conn_str = (
    r'DRIVER={Microsoft Access Driver (*.mdb, *.accdb)};'
    r'DBQ=..\01_database\PL-182 HUSOW.mdb;'  # ścieżka do bazy danych, do zmiany jeśli potrzeba
    )

cnxn = pyodbc.connect(conn_str)
crsr = cnxn.cursor()

# dane dzienne z bazy jak trzeba popraw kwerendę
qra = "Select " \
      "`Station (text)`, `Station (value)`, " \
      "`Track`, `Bin`, " \
      "`Descriptor`, `Description1`, `Description2`, `Comment`, " \
      "`Survey Time (Local)`, `Survey Mode (text)`, `Surveyor` " \
      "From [POSTPLOT] " \
      "Where " \
      "(datediff ('d', `Survey Time (Local)`,Now()) = 1) And " \
      "(" \
      "(`Station (value)` > 0 and `Station (text)` Not Like '88%') " \
      "OR `Station (text)` Like 'cp%' " \
      "OR `Station (text)` Like '?88%'" \
      ")  " \
      "Order By `Surveyor`,`Survey Time (Local)`"

crsr.execute(qra)
dane = crsr.fetchall()
print(type(dane), len(dane))
cnxn.commit()
cnxn.close()


print('Otwieram plik excel')

wb = openpyxl.load_workbook(filename='2021_PL182_3DRaport_Geodezja.xlsm', read_only=False, keep_vba=True)
ws = wb.worksheets[1]

# Czyszczenie arkusza z danych z dnia wcześniejszego
for row in ws.iter_rows(min_row=2, max_col=11):
    for cell in row:
        cell.value = None


# wypełnienie wierszy wartościami z danych
row = 1
for ro in dane:
    row += 1
    ws['A' + str(row)] = ro[0]
    ws['B' + str(row)] = ro[1]
    ws['C' + str(row)] = ro[2]
    ws['D' + str(row)] = ro[3]
    ws['E' + str(row)] = ro[4]
    ws['F' + str(row)] = ro[5]
    ws['G' + str(row)] = ro[6]
    ws['H' + str(row)] = ro[7]
    ws['I' + str(row)] = ro[8]
    ws['J' + str(row)] = ro[9]
    ws['K' + str(row)] = ro[10]



ws.title = '20210809'
# Zmiana nazwy arkusza na prawidłową
#    ws.title = str(date.today().strftime('%Y%m%d'))

# Zapisanie zmian (jak narazie w alternatywnym pliku)
#wb.save('test.xlsm')
print(ws.title)


#exit()

#print(std)
'''
ws.append(['Station (text)', 'Station (value)', 'Track', 'Bin', 'Descriptor',
           'Description1', 'Description2', 'Comment', 'Survey Time (Local)',
           'Survey Mode (text)', 'Surveyor'])
'''
#for row in std:
#    ws.append(item for item in row)

wb.save("test.xlsm")
# wb.save(f"../08_raport_geodezyjny/dniowki/{str(date.today().strftime('%Y%m%d'))}.xlsx")

print('Gotowe')
input("Enter by zakończyć")
