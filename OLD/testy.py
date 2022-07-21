from openpyxl import Workbook
from openpyxl.styles import Alignment

wb = Workbook()
ws = wb.active

#【A】-----------------------------------------------------------
# Merge multiple cells by merge_cells() method (A1 address specified)
ws.merge_cells('B2:F4')

# Acquisition of the Cell object (B2) in the upper left corner of the merged cell area
top_left_cell = ws['B2']

# Decorate entire merged cell
# The entire combined cell is decorated by setting the style formatting on the top left Cell object (B2)
# Note that specifying cells other than the upper left, such as B3, C2, etc., will result in an error.
# Also, borders, etc., cannot be set at once, but must be set for each individual cell.

top_left_cell.value = "Merged Cell!"
top_left_cell.alignment = Alignment(horizontal="center", vertical="center")

#【B】-----------------------------------------------------------
# Merge multiple cells by merge_cells() method (specify number of matrices)
ws.merge_cells(start_row= 7, end_row=9, start_column=2, end_column=3)


wb.save('sample_Merged-Cell.xlsx')