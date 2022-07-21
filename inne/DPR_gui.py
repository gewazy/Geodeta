import pyodbc
import openpyxl
from datetime import date, timedelta
from time import sleep
import kury
import init

f = open('zwrot.txt', 'w', encoding='utf-8')

f.write('Generowania raportu DPR.\n\tver3.beta')

f.write(f"-dane z dnia: {str(date.today().strftime('%Y%m%d'))}")

# łączenie z bazą
cnxn = pyodbc.connect(init.CONN_STR)  # łączenie z bazą danych
crsr = cnxn.cursor()

# zbieranie danch do wypełnienia raportu
vib = len(crsr.execute(kury.VIB).fetchall())
xr = len(crsr.execute(kury.XR).fetchall())
xt = len(crsr.execute(kury.XT).fetchall())
skip = len(crsr.execute(kury.SKIP_S).fetchall())
qc_r = len(crsr.execute(kury.QC_R).fetchall())
qc_s = len(crsr.execute(kury.QC_S).fetchall())
paterny = len(crsr.execute(kury.PATERNY).fetchall())


tycz_r = crsr.execute(kury.TYCZ_R).fetchall()
tycz_s = crsr.execute(kury.TYCZ_S).fetchall()

zm_r = crsr.execute(kury.ZM_R).fetchall()
zm_s = crsr.execute(kury.ZM_S).fetchall()

re_s = crsr.execute(kury.RE_S).fetchall()
re_r = crsr.execute(kury.RE_R).fetchall()
otg = crsr.execute(kury.OTG).fetchall()

crsr.close()
cnxn.close()


# Przygotowanie pliku Excell
# print('Otwieram DPR\n')

wb = openpyxl.load_workbook(init.DPR[0])

# print('-Tworzę nową zakładkę')
source = wb[wb.sheetnames[-2]]
target = wb.copy_worksheet(source)
wb.move_sheet(target, offset=-1)

# print('-Zmieniam nazwe zakładki')
target.title = str((date.today() + timedelta(days=1)).strftime('%Y%m%d'))
target.sheet_view.zoomScale = 75
target.sheet_view.view = "pageBreakPreview"

# zamiana formuł w pliku
# print('-Edytuję formuły')
a = f"'{str((date.today() - timedelta(days=1)).strftime('%Y%m%d'))}'"
b = f"'{source.title}'"
for row in target.iter_rows(min_row=1, max_col=15, max_row=133):
    for cell in row:
        if type(cell.value) is str:
            cell.value = cell.value.replace(a, b)

# ustawienie arkusza aktywnego
# print('\nPrzechodzę na arkusz z dnia dzisiejszego')
target = wb[wb.sheetnames[-3]]


# wypełnianie raportu
f.write("-Wypełniłem arkusz\n")

f.write(f'--Wibratory: {vib}\n')
target['E132'] = vib

f.write(f'--Wiercenie ręczne: {xr}\n')
target['F132'] = xr

f.write(f'--Wiercenie traktorem: {xt}\n')
target['G132'] = xt

f.write(f'--W tym patterny: {paterny}\n')
target['H134'] = paterny

f.write(f'--Skipy: {skip}\n')
target['K132'] = skip

f.write(f'--QC R: {qc_r}\n')
if qc_r != 0:
    target['G53'] = qc_r

f.write(f'--QC S: {qc_s}\n')
if qc_s != 0:
    target['O53'] = qc_s

target['C132'] = init.GEODETA

licz_bry = []  # lista brygadzistów do liczenia liczby brygad

row = 13
f.write('\nTyczenie punktów odbioru: \n')
for ro in tycz_r:
    row += 1
    f.write(f"{ro}\n")
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    target['D' + str(row)] = ro[3]
    try:
        licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue

row = 13
f.write('\nTyczenie punktów wzbudzania: \n')
for ro in tycz_s:
    row += 1
    f.write(f"{ro}\n")
    target['I' + str(row)] = ro[0]
    target['J' + str(row)] = ro[1]
    target['K' + str(row)] = ro[2]
    target['L' + str(row)] = ro[3]
    try:
        licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue

row = 41
f.write('\nDomierzanie/niwelacja punktów odbioru: \n')
for ro in re_r:
    row += 1
    f.write(f"{ro}\n")
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    target['D' + str(row)] = ro[3]
    try:
        licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue

row = 41
f.write('\nDomierzanie/niwelacja punktów wzbudzania: \n')
for ro in re_s:
    row += 1
    f.write(f"{ro}\n")
    target['I' + str(row)] = ro[0]
    target['J' + str(row)] = ro[1]
    target['K' + str(row)] = ro[2]
    target['L' + str(row)] = ro[3]
    try:
        licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue

# rozróżnienie zmian od niwelacji/domierzania /geofony/
for r1 in zm_r:
    for r2 in re_r:
        if r1[2] == r2[2]:
            r1[3] -= r2[3]

row = 57
f.write('\nZmiany punktów odbioru: \n')
for ro in zm_r:
    if ro[3] > 0:
        row += 1
        f.write(f"{ro}\n")
        target['A' + str(row)] = ro[0]
        target['B' + str(row)] = ro[1]
        target['C' + str(row)] = ro[2]
        target['D' + str(row)] = ro[3]
        try:
            licz_bry.append((ro[2].split())[1])
        except IndexError:
            continue

# rozróżnienie zmian od niwelacji, domierzania /pkt strzałowe/
for r1 in zm_s:
    for r2 in re_s:
        if r1[2] == r2[2]:
            r1[3] -= r2[3]

row = 57
f.write('\nZmiany punktów wzbudzania: \n')
for ro in zm_s:
    if ro[3] > 0:
        row += 1
        f.write(f"{ro}\n")
        target['I' + str(row)] = ro[0]
        target['J' + str(row)] = ro[1]
        target['K' + str(row)] = ro[2]
        target['L' + str(row)] = ro[3]
        try:
            licz_bry.append((ro[2].split())[1])
        except IndexError:
            continue

row = 95
f.write('\nPomiar otworów głębokich: \n')
for ro in otg:
    row += 1
    f.write(f"{ro}\n")
    target['A' + str(row)] = ro[0]
    target['B' + str(row)] = ro[1]
    target['C' + str(row)] = ro[2]
    target['D' + str(row)] = ro[3]
    try:
        licz_bry.append((ro[2].split())[1])
    except IndexError:
        continue

f.write(f'\n\nPracowało {len(set(licz_bry))} brygad\n')
for num, geodeta in enumerate(sorted(set(licz_bry))):
    f.write(f"{str(num + 1)}. {geodeta}")

target['N132'] = len(set(licz_bry))
# target['B128'] = input('\nWprowadź komentarz:\n')

wb.save(init.DPR[0])

f.close()
